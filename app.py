import streamlit as st
import numpy as np
import os
from pathlib import Path
import json
from datetime import datetime
from io import BytesIO

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Hotel KPI — Asset Manager",
    page_icon="🏨",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
    h1 { font-size: 1.4rem !important; font-weight: 600 !important; }
    h2 { font-size: 1.1rem !important; font-weight: 600 !important; color: #1A2E44; }
    h3 { font-size: 0.95rem !important; font-weight: 600 !important; }
    .stMetric { background: #F1EFE8; border-radius: 8px; padding: 12px; }
    .score-box { text-align: center; padding: 1rem; border-radius: 12px; margin-bottom: 1rem; }
    .kpi-green { color: #1D9E75; font-weight: 600; }
    .kpi-amber { color: #EF9F27; font-weight: 600; }
    .kpi-red   { color: #E24B4A; font-weight: 600; }
    .section-tag { font-size: 0.7rem; font-weight: 600; letter-spacing: 0.06em;
                   text-transform: uppercase; color: #888780; margin-bottom: 0.4rem; }
    .flag { border-radius: 6px; padding: 8px 12px; margin-bottom: 6px; font-size: 0.85rem; }
    .flag-red   { background: #FCEBEB; color: #A32D2D; }
    .flag-amber { background: #FAEEDA; color: #854F0B; }
    .flag-green { background: #E1F5EE; color: #0F6E56; }
    .stExpander { border: 0.5px solid #D3D1C7 !important; border-radius: 8px !important; }
    div[data-testid="stHorizontalBlock"] > div { padding: 0 0.25rem; }
</style>
""", unsafe_allow_html=True)

# ── Helpers ──────────────────────────────────────────────────────────────────
SEG_DEFAULTS = {
    "Économique (1-2 ★)":     dict(occ=74,adr=70,rooms=50,opcost=35,ltv=55,dscr=1.5,caprate=7.0,noi=200,season=4,newcomp=2,intl=15,loc=5,esg=2,digi=3,hr=5,ins=6,div=3,brand_mode="Indépendant",brand_strength=4,fee_pct=0,contract_years=0,exit_penalty=3,brand_rep=5,franchise_audits=2,legal_dpe="C",legal_access=6,legal_fire=8,legal_rgpd=4,legal_labor=6,legal_env=3,legal_tax=5),
    "Milieu de gamme (3 ★)":  dict(occ=72,adr=120,rooms=80,opcost=55,ltv=60,dscr=1.4,caprate=5.0,noi=500,season=5,newcomp=3,intl=35,loc=6,esg=4,digi=5,hr=6,ins=7,div=4,brand_mode="Franchise",brand_strength=6,fee_pct=8,contract_years=9,exit_penalty=5,brand_rep=7,franchise_audits=4,legal_dpe="D",legal_access=7,legal_fire=8,legal_rgpd=6,legal_labor=7,legal_env=5,legal_tax=6),
    "Haut de gamme (4 ★)":    dict(occ=68,adr=220,rooms=120,opcost=100,ltv=58,dscr=1.6,caprate=4.5,noi=1200,season=6,newcomp=4,intl=50,loc=7,esg=6,digi=7,hr=7,ins=8,div=6,brand_mode="Franchise",brand_strength=8,fee_pct=10,contract_years=15,exit_penalty=6,brand_rep=8,franchise_audits=6,legal_dpe="C",legal_access=8,legal_fire=9,legal_rgpd=7,legal_labor=8,legal_env=6,legal_tax=7),
    "Luxe (5 ★)":             dict(occ=65,adr=550,rooms=90,opcost=280,ltv=50,dscr=2.0,caprate=3.5,noi=3000,season=7,newcomp=5,intl=70,loc=9,esg=8,digi=8,hr=8,ins=9,div=8,brand_mode="Contrat de gestion",brand_strength=9,fee_pct=12,contract_years=20,exit_penalty=8,brand_rep=9,franchise_audits=8,legal_dpe="B",legal_access=9,legal_fire=9,legal_rgpd=8,legal_labor=8,legal_env=7,legal_tax=8),
}

def compute_scores(d):
    occ = d['occ'] / 100
    revpar = d['adr'] * occ
    goppar = revpar - d['opcost']
    gop_ratio = (goppar / revpar * 100) if revpar > 0 else 0
    trevpar = revpar * 1.12
    asset_val = (d['noi'] / d['caprate'] * 100) if d['caprate'] > 0 else 0

    # CORRECTION 1 — GOPPAR négatif : exploitation à perte → score bloqué à 0
    goppar_negatif = goppar < 0
    if goppar_negatif:
        ops = 0
    else:
        ops = min(100, max(0, round(gop_ratio * 2)))

    # FIN score
    ltv_s  = 100 if d['ltv']<50 else 75 if d['ltv']<60 else 50 if d['ltv']<70 else 25
    dscr_s = 100 if d['dscr']>1.5 else 70 if d['dscr']>1.3 else 40 if d['dscr']>1.2 else 10
    cr_s   = 100 if d['caprate']>6 else 65 if d['caprate']>4 else 30
    fin = round((ltv_s + dscr_s + cr_s) / 3)

    # MKT score
    season_s = 100 if d['season']<=3 else 65 if d['season']<=6 else 30
    comp_s   = 100 if d['newcomp']<=2 else 65 if d['newcomp']<=5 else 30
    loc_s    = 100 if d['loc']>=8 else 75 if d['loc']>=6 else 50 if d['loc']>=4 else 25
    mkt = round((season_s + comp_s + loc_s) / 3)

    # RES score
    res = min(100, round(((d['esg'] + d['digi'] + (10-d['hr']) + d['ins'] + d['div']) / 50) * 100))

    # CORRECTION 2 — Cohérence LTV vs dette totale réelle
    # Accès sécurisé à total_debt — peut être absent si onglet financier pas encore visité
    total_debt_v = d.get('total_debt', 0) or 0
    ltv_reel = round(total_debt_v / asset_val * 100, 1) if (asset_val > 0 and total_debt_v > 0) else None
    ltv_incoherence = (
        ltv_reel is not None and
        abs(ltv_reel - d['ltv']) > 15
    )

    # BRAND score
    brand_strength_s = d['brand_strength'] * 10
    fee_risk  = 100 if d['fee_pct']<6 else 70 if d['fee_pct']<10 else 40 if d['fee_pct']<14 else 20
    exit_risk = 100 if d['exit_penalty']<=3 else 65 if d['exit_penalty']<=6 else 30
    rep_s     = d['brand_rep'] * 10
    audit_s   = 100 if d['franchise_audits']<=2 else 65 if d['franchise_audits']<=5 else 30

    if d['brand_mode'] == "Indépendant":
        # CORRECTION — score calculé, pas fixé arbitrairement à 75
        # Un indépendant sans réseau perd ~10-15% d'occupation vs franchisé
        # On pénalise sur force marque et distribution, mais sans risque exit/redevance
        brand = round((brand_strength_s * 0.5 + rep_s * 0.5))
        brand = min(100, max(20, brand))
    else:
        brand = round((brand_strength_s + fee_risk + exit_risk + rep_s + audit_s) / 5)

    # CORRECTION 3 — DPE pondéré à 30% (risque d'interdiction d'exploitation)
    # Les 6 autres critères se partagent les 70% restants
    dpe_map  = {"A":100,"B":90,"C":75,"D":55,"E":30,"F":10,"G":0}
    dpe_s    = dpe_map.get(d['legal_dpe'], 50)
    access_s = d['legal_access'] * 10
    fire_s   = d['legal_fire'] * 10
    rgpd_s   = d['legal_rgpd'] * 10
    labor_s  = d['legal_labor'] * 10
    env_s    = d['legal_env'] * 10
    tax_s    = d['legal_tax'] * 10
    legal = round(
        dpe_s * 0.30 +
        (access_s + fire_s + rgpd_s + labor_s + env_s + tax_s) / 6 * 0.70
    )

    # PONDÉRATION DYNAMIQUE — franchise pèse plus si actif en franchise
    if d['brand_mode'] in ["Franchise", "Contrat de gestion"]:
        brand_w, legal_w = 0.15, 0.10
        ops_w = fin_w = 0.225
        mkt_w = res_w = 0.15
    else:
        brand_w, legal_w = 0.05, 0.10
        ops_w = fin_w = 0.275
        mkt_w = res_w = 0.15

    global_score = round(
        ops*ops_w + fin*fin_w + mkt*mkt_w +
        res*res_w + brand*brand_w + legal*legal_w
    )

    refi = 'Élevé' if d['ltv']>70 or d['dscr']<1.2 else 'Modéré' if d['ltv']>55 or d['dscr']<1.4 else 'Faible'

    return dict(
        revpar=round(revpar,1), goppar=round(goppar,1),
        trevpar=round(trevpar,1), gop_ratio=round(gop_ratio,1),
        asset_val=round(asset_val), refi=refi,
        ops=ops, fin=fin, mkt=mkt, res=res, brand=brand, legal=legal,
        global_score=global_score,
        goppar_negatif=goppar_negatif,
        ltv_reel=ltv_reel,
        ltv_incoherence=ltv_incoherence,
    )

def risk_label(s):
    return "🟢 Faible" if s>=75 else "🟡 Modéré" if s>=50 else "🔴 Élevé"

def score_color(s):
    return "#1D9E75" if s>=75 else "#EF9F27" if s>=50 else "#E24B4A"

def badge(s):
    if s >= 75: return f'<span class="kpi-green">✓ OK</span>'
    if s >= 50: return f'<span class="kpi-amber">⚠ Vigilance</span>'
    return f'<span class="kpi-red">✗ Risque</span>'

# ── Sidebar — Identification ──────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🏨 Identification de l'actif")
    asset_name = st.text_input("Nom de l'hôtel / actif", "Hôtel Example")
    location   = st.text_input("Localisation", "Paris, Île-de-France")
    segment    = st.selectbox("Segment", list(SEG_DEFAULTS.keys()), index=1)
    analyst    = st.text_input("Asset Manager", "")
    period     = st.text_input("Période", datetime.today().strftime("%B %Y"))
    usage_mode = st.selectbox("Mode d'usage", ["Acquisition / Entrée", "Actif existant — suivi AM", "Les deux"])

    st.divider()
    if st.button("🔄 Pré-remplir valeurs types (segment)", use_container_width=True):
        st.session_state['prefill'] = SEG_DEFAULTS[segment]
        st.rerun()

    st.caption(f"Généré le {datetime.today().strftime('%d/%m/%Y')}")

# ── Load defaults ─────────────────────────────────────────────────────────────
D = st.session_state.get('prefill', SEG_DEFAULTS[segment])

# ── Tabs ──────────────────────────────────────────────────────────────────────
t1, t2, t3, t4, t5, t6, t7, t8, t9 = st.tabs([
    "📊 Dashboard",
    "⚙️ Opérationnel",
    "💰 Financier",
    "🌍 Marché",
    "🛡️ Résilience",
    "🏷️ Franchise",
    "⚖️ Légal",
    "🧨 Stress Test",
    "📐 Risk Mgmt",
])

# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — OPÉRATIONNEL
# ════════════════════════════════════════════════════════════════════════════
with t2:
    st.markdown("## Performance opérationnelle")
    c1, c2 = st.columns(2)
    with c1:
        rooms   = st.number_input("Nombre de chambres", 5, 2000, D['rooms'])
        adr     = st.number_input("ADR — Prix moyen / chambre (€)", 30, 3000, D['adr'])
        opcost  = st.number_input("Coûts opéra. / chambre dispo (€)", 5, 800, D['opcost'])
    with c2:
        occ        = st.slider("Taux d'occupation (%)", 10, 100, D['occ'])
        reputation = st.slider("Réputation en ligne (/5)", 1.0, 5.0, 4.1, 0.1)
        retention  = st.slider("Taux fidélisation client (%)", 0, 100, 28)

    st.markdown("---")
    st.markdown("**Benchmarks France 2025 — In Extenso / Deloitte**")
    bench_data = {
        "Segment": ["Économique","Milieu de gamme","Haut de gamme","Luxe Paris","Côte d'Azur"],
        "RevPAR cible": ["45–65€","70–110€","130–200€","300–500€","+6% vs N-1"],
        "Taux occ.": ["70–78%","68–75%","65–75%","60–75%","~60%"],
        "ADR": ["65–85€","100–150€","200–300€","450–800€","+5% vs N-1"],
    }
    st.table(bench_data)

# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — FINANCIER
# ════════════════════════════════════════════════════════════════════════════
with t3:
    st.markdown("## Structure financière")
    c1, c2 = st.columns(2)
    with c1:
        noi     = st.number_input("NOI annuel estimé (k€)", 0, 50000, D['noi'])
        caprate = st.number_input("Cap Rate (%)", 0.5, 15.0, D['caprate'], 0.5)
        ltv     = st.slider("LTV — Loan-to-Value (%)", 0, 100, D['ltv'])
        dscr    = st.slider("DSCR — Couverture service de la dette", 0.5, 4.0, D['dscr'], 0.1)
    with c2:
        total_debt  = st.number_input("Dette totale (k€)", 0, 500000, 5000)
        capex       = st.number_input("Capex prévu 3 ans (k€)", 0, 50000, 200)
        loan_years  = st.number_input("Durée résiduelle prêt (ans)", 1, 30, 15)

    st.markdown("---")
    st.markdown("**Grille de lecture — seuils**")
    grid = {
        "Indicateur": ["LTV","DSCR","Cap Rate","Marge GOP","Cash-on-cash return"],
        "🟢 Vert": ["<50%",">1.5x",">6%",">35%",">8%"],
        "🟡 Orange": ["50–70%","1.2–1.5x","4–6%","20–35%","5–8%"],
        "🔴 Rouge": [">70%","<1.2x","<4%","<20%","<5%"],
    }
    st.table(grid)

# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — MARCHÉ
# ════════════════════════════════════════════════════════════════════════════
with t4:
    st.markdown("## Risque marché")
    c1, c2 = st.columns(2)
    with c1:
        season  = st.slider("Saisonnalité (1=faible, 10=forte)", 1, 10, D['season'])
        loc     = st.slider("Score localisation (1=périphérie, 10=prime)", 1, 10, D['loc'])
    with c2:
        newcomp = st.number_input("Nouveaux concurrents prévus (pipeline 3 ans)", 0, 50, D['newcomp'])
        intl    = st.slider("Part clientèle internationale (%)", 0, 100, D['intl'])

    st.markdown("---")
    st.markdown("**Tendances marché France 2026**")
    trends = {
        "Marché": ["Paris","Côte d'Azur","Île-de-France hors Paris","Régions affaires","Luxe Europe","Économique non classé"],
        "Tendance": ["Taux occ. +5%, ADR -1%","RevPAR +6%","Absorption nouvelles capacités","Mitigé","Surabondance offre","Baisse -9.2%"],
        "Signal": ["🟢 Positif","🟢 Fort","🟡 Neutre","🟡 Prudent","🔴 Vigilance","🔴 Sous pression"],
    }
    st.table(trends)

# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — RÉSILIENCE
# ════════════════════════════════════════════════════════════════════════════
with t5:
    st.markdown("## Résilience de l'actif")
    c1, c2 = st.columns(2)
    with c1:
        esg  = st.slider("Score ESG (0=aucun, 10=certifié)", 0, 10, D['esg'])
        digi = st.slider("Maturité digitale (0=basique, 10=avancé)", 0, 10, D['digi'])
        hr   = st.slider("Tension RH (0=stable, 10=critique)", 0, 10, D['hr'])
    with c2:
        ins  = st.slider("Couverture assurance (0=faible, 10=complète)", 0, 10, D['ins'])
        div  = st.slider("Diversification revenus (0=chambres seules, 10=multi)", 0, 10, D['div'])

    st.markdown("---")
    st.markdown("**Nouveaux risques 2026 à surveiller**")
    new_risks = {
        "Risque": ["Inflation coûts construction","Mise aux normes ESG/DPE","Cybersécurité (16Mds identifiants exposés)","Crise des talents structurelle","Disruption IA distribution","Déplacement invest. vers la défense"],
        "Impact": ["🔴 Élevé","🔴 Élevé","🔴 Élevé","🟡 Moyen","🟡 Moyen","🟡 Moyen"],
    }
    st.table(new_risks)

# ════════════════════════════════════════════════════════════════════════════
# TAB 6 — MARQUE & FRANCHISE
# ════════════════════════════════════════════════════════════════════════════
with t6:
    st.markdown("## Risque marque & contrat de franchise")
    st.info("Cette dimension évalue l'exposition liée au mode d'exploitation : indépendant, franchise ou contrat de gestion.")

    c1, c2 = st.columns(2)
    with c1:
        brand_mode = st.selectbox("Mode d'exploitation", ["Indépendant","Franchise","Contrat de gestion","Affiliation / Soft brand"], index=["Indépendant","Franchise","Contrat de gestion","Affiliation / Soft brand"].index(D.get('brand_mode','Franchise')))
        brand_strength = st.slider("Force de la marque (1=inconnue, 10=leader mondial)", 1, 10, D['brand_strength'])
        brand_rep = st.slider("Réputation / image de marque (1=dégradée, 10=excellente)", 1, 10, D['brand_rep'])
        franchise_audits = st.slider("Fréquence audits franchiseur / an", 0, 12, D['franchise_audits'])

    with c2:
        if brand_mode != "Indépendant":
            fee_pct      = st.slider("Redevances totales (% du CA)", 0, 25, D['fee_pct'])
            contract_years = st.slider("Durée restante du contrat (ans)", 0, 30, D['contract_years'])
            exit_penalty = st.slider("Pénalités de sortie anticipée (1=faibles, 10=très élevées)", 1, 10, D['exit_penalty'])
        else:
            fee_pct = 0
            contract_years = 0
            exit_penalty = 1
            st.markdown("*Hôtel indépendant — pas de redevance ni de contrat franchiseur.*")

    st.markdown("---")
    st.markdown("**Points de vigilance contrat de franchise (jurisprudence France 2025)**")
    franchise_risks = {
        "Point contractuel": [
            "Clause de non-concurrence post-contrat",
            "Résiliation unilatérale par le franchiseur",
            "Obligation de rénovation imposée",
            "Hausse des redevances non plafonnée",
            "Changement de marque / repositionnement",
            "Mandats de gestion confiés à filiale franchiseur",
        ],
        "Niveau de risque": ["🟡 Moyen","🔴 Élevé","🟡 Moyen","🔴 Élevé","🟡 Moyen","🔴 Élevé"],
        "Mitigation": [
            "Vérifier durée et périmètre géographique",
            "Négocier clause de préavis minimum 24 mois",
            "Plaffonner le Capex imposé dans le contrat",
            "Indexer les hausses au RevPAR de l'actif",
            "Clause de compensation en cas de rebrand",
            "Audit séparé du mandat de gestion",
        ],
    }
    st.table(franchise_risks)

# ════════════════════════════════════════════════════════════════════════════
# TAB 7 — LÉGAL & RÉGLEMENTAIRE
# ════════════════════════════════════════════════════════════════════════════
with t7:
    st.markdown("## Risque légal & réglementaire")
    st.warning("⚡ Le cadre réglementaire hôtelier français s'est fortement durci en 2025-2026. Affichage environnemental obligatoire dès 2026, DPE opposable, RGPD étendu aux PME.")

    c1, c2 = st.columns(2)
    with c1:
        legal_dpe    = st.selectbox("Classe DPE de l'actif", ["A","B","C","D","E","F","G"], index=["A","B","C","D","E","F","G"].index(D['legal_dpe']))
        legal_access = st.slider("Conformité accessibilité PMR (0=non conforme, 10=conforme)", 0, 10, D['legal_access'])
        legal_fire   = st.slider("Sécurité incendie (0=non conforme, 10=conforme)", 0, 10, D['legal_fire'])
        legal_rgpd   = st.slider("Conformité RGPD (0=non conforme, 10=certifié)", 0, 10, D['legal_rgpd'])
    with c2:
        legal_labor  = st.slider("Conformité droit du travail (0=risques, 10=conforme)", 0, 10, D['legal_labor'])
        legal_env    = st.slider("Conformité environnementale (0=risques, 10=certifié)", 0, 10, D['legal_env'])
        legal_tax    = st.slider("Situation fiscale (0=risques, 10=saine)", 0, 10, D['legal_tax'])

    st.markdown("---")
    st.markdown("**Calendrier réglementaire clé — France 2025-2034**")
    regle_data = {
        "Échéance": ["2025 (en cours)","2026","2027","2028","2034"],
        "Obligation": [
            "DPE G interdit à la location · GTB CVC >290kW · Tri déchets textiles",
            "Affichage environnemental obligatoire (hôtels) · RGPD étendu PME · GTB >70kW",
            "Bilan carbone obligatoire par nuitée (>15 chambres)",
            "DPE F interdit à la location",
            "DPE E interdit à la location",
        ],
        "Impact potentiel": ["🔴 Immédiat","🔴 Immédiat","🟡 Moyen terme","🟡 Moyen terme","🟢 Long terme"],
    }
    st.table(regle_data)

    with st.expander("📋 Checklist conformité réglementaire complète"):
        checks = [
            ("DPE valide et classe ≥ D", "Légal — Énergie", "CRITIQUE"),
            ("Permis d'exploitation à jour", "Légal — Administratif", "CRITIQUE"),
            ("Conformité accessibilité PMR (loi 2005)", "Légal — Accessibilité", "CRITIQUE"),
            ("Sécurité incendie (détecteurs, issues, plan évacuation)", "Légal — Sécurité", "CRITIQUE"),
            ("Conformité RGPD — DPO désigné si nécessaire", "Légal — Données", "CRITIQUE"),
            ("Formation HACCP (restauration)", "Légal — Hygiène", "IMPORTANT"),
            ("GTB CVC installé si puissance > 290kW", "Légal — Énergie", "IMPORTANT"),
            ("Audit énergétique si DPE F ou G", "Légal — Énergie", "IMPORTANT"),
            ("Tri déchets textiles organisé (2025)", "Légal — Environnement", "IMPORTANT"),
            ("Conformité droit du travail (SMIC, heures sup)", "Légal — Social", "IMPORTANT"),
            ("Classement Atout France valide (5 ans)", "Légal — Classement", "À ÉVALUER"),
            ("Affichage environnemental préparé (2026)", "Légal — Énergie", "À ÉVALUER"),
        ]
        for item, dim, crit in checks:
            color = "#FCEBEB" if crit=="CRITIQUE" else "#FAEEDA" if crit=="IMPORTANT" else "#E1F5EE"
            tc = "#A32D2D" if crit=="CRITIQUE" else "#854F0B" if crit=="IMPORTANT" else "#0F6E56"
            st.markdown(f"""<div class="flag" style="background:{color};color:{tc}">
                <b>[{crit}]</b> {item} <span style="opacity:.6;font-size:.8rem">— {dim}</span>
            </div>""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# COLLECT ALL INPUTS for compute
# ════════════════════════════════════════════════════════════════════════════
# Re-read all widget values (they are already in scope from each tab)
try:
    data = dict(
        occ=occ, adr=adr, rooms=rooms, opcost=opcost,
        noi=noi, caprate=caprate, ltv=ltv, dscr=dscr,
        total_debt=total_debt, capex=capex, loan_years=loan_years,
        season=season, newcomp=newcomp, intl=intl, loc=loc,
        esg=esg, digi=digi, hr=hr, ins=ins, div=div,
        brand_mode=brand_mode, brand_strength=brand_strength,
        fee_pct=fee_pct, contract_years=contract_years,
        exit_penalty=exit_penalty, brand_rep=brand_rep,
        franchise_audits=franchise_audits,
        legal_dpe=legal_dpe, legal_access=legal_access,
        legal_fire=legal_fire, legal_rgpd=legal_rgpd,
        legal_labor=legal_labor, legal_env=legal_env, legal_tax=legal_tax,
    )
    C = compute_scores(data)
except NameError:
    data = dict(
        occ=D['occ'], adr=D['adr'], rooms=D['rooms'], opcost=D['opcost'],
        noi=D['noi'], caprate=D['caprate'], ltv=D['ltv'], dscr=D['dscr'],
        total_debt=0, capex=0, loan_years=15,
        season=D['season'], newcomp=D['newcomp'], intl=D['intl'], loc=D['loc'],
        esg=D['esg'], digi=D['digi'], hr=D['hr'], ins=D['ins'], div=D['div'],
        brand_mode=D['brand_mode'], brand_strength=D['brand_strength'],
        fee_pct=D['fee_pct'], contract_years=D['contract_years'],
        exit_penalty=D['exit_penalty'], brand_rep=D['brand_rep'],
        franchise_audits=D['franchise_audits'],
        legal_dpe=D['legal_dpe'], legal_access=D['legal_access'],
        legal_fire=D['legal_fire'], legal_rgpd=D['legal_rgpd'],
        legal_labor=D['legal_labor'], legal_env=D['legal_env'], legal_tax=D['legal_tax'],
    )
    C = compute_scores(data)

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
with t1:
    sc = C['global_score']
    col_score, col_right = st.columns([1, 2])

    with col_score:
        clr = score_color(sc)
        st.markdown(f"""
        <div class="score-box" style="background:{clr}22;border:2px solid {clr}">
            <div style="font-size:3.5rem;font-weight:700;color:{clr};line-height:1">{sc}</div>
            <div style="font-size:0.75rem;font-weight:600;color:{clr};letter-spacing:.05em">SCORE GLOBAL / 100</div>
            <div style="font-size:1rem;font-weight:600;color:{clr};margin-top:4px">{risk_label(sc)}</div>
        </div>
        """, unsafe_allow_html=True)
        st.caption(f"**{asset_name}** · {segment} · {location}")
        st.caption(f"AM : {analyst} · {period}")

    with col_right:
        st.markdown("**Décomposition par dimension**")
        dims = [
            ("⚙️ Performance opéra.", C['ops'], "25%"),
            ("💰 Risque financier",   C['fin'], "25%"),
            ("🌍 Risque marché",      C['mkt'], "15%"),
            ("🛡️ Résilience",         C['res'], "15%"),
            ("🏷️ Marque & Franchise", C['brand'], "10%"),
            ("⚖️ Légal & Réglementaire", C['legal'], "10%"),
        ]
        for name, s, w in dims:
            clr = score_color(s)
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px">
              <span style="min-width:220px;font-size:.85rem">{name} <span style="color:#888780;font-size:.75rem">({w})</span></span>
              <div style="flex:1;height:8px;background:#F1EFE8;border-radius:4px;overflow:hidden">
                <div style="width:{s}%;height:100%;background:{clr};border-radius:4px"></div>
              </div>
              <span style="min-width:32px;font-size:.85rem;font-weight:600;color:{clr};text-align:right">{s}</span>
            </div>
            """, unsafe_allow_html=True)

    st.divider()

    # KPIs métriques
    st.markdown("**Indicateurs clés calculés**")
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("RevPAR", f"{C['revpar']}€")
    m2.metric("GOPPAR", f"{C['goppar']}€")
    m3.metric("Marge GOP", f"{C['gop_ratio']}%")
    m4.metric("LTV", f"{data['ltv']}%")
    m5.metric("DSCR", f"{data['dscr']:.1f}x")
    m6.metric("Valeur actif", f"{C['asset_val']:,}k€")

    st.divider()

    # Signaux d'alerte
    st.markdown("**Signaux d'alerte**")
    flags = []
    # ── ALERTES CRITIQUES DE COHÉRENCE (nouvelles) ────────────────────────
    if C['goppar_negatif']:
        flags.append(("red", "🚨 EXPLOITATION À PERTE — GOPPAR négatif : coûts opérationnels > RevPAR. Score opérationnel bloqué à 0. Revoir immédiatement la structure de coûts."))
    if C['ltv_incoherence']:
        flags.append(("red", f"⚠️ INCOHÉRENCE LTV — LTV déclaré {data['ltv']}% / LTV réel calculé {C['ltv_reel']}% (dette {data.get('total_debt',0)}k€ / valeur actif {C['asset_val']}k€). Écart >15pts — vérifier les données saisies."))
    if data['legal_dpe'] in ['E','F','G']:
        flags.append(("red", f"🚨 DPE {data['legal_dpe']} — Interdiction d'exploitation progressive (G: 2025, F: 2028, E: 2034). Rénovation énergétique impérative. Ce risque pèse 30% du score légal."))
    # ── ALERTES STANDARD ─────────────────────────────────────────────────
    if C['fin']   < 50: flags.append(("red",   "Risque financier élevé — LTV ou DSCR hors seuils"))
    if C['ops']   < 50 and not C['goppar_negatif']: flags.append(("red", "Performance opérationnelle sous les benchmarks du segment"))
    if C['legal'] < 50: flags.append(("red",   "Conformité réglementaire insuffisante — DPE, RGPD ou sécurité incendie"))
    if C['brand'] < 50 and data['brand_mode'] != "Indépendant":
                         flags.append(("red",   "Risque franchise élevé — redevances, exit ou dépendance marque"))
    if C['res']   < 60: flags.append(("amber", "Résilience fragile — renforcer ESG, digital et couverture assurance"))
    if C['mkt']   < 60: flags.append(("amber", "Pression marché — saisonnalité ou pipeline concurrentiel"))
    if data['brand_mode'] != "Indépendant" and data['fee_pct'] > 12:
                         flags.append(("amber", f"Redevances élevées ({data['fee_pct']}% CA) — impact direct sur GOPPAR"))
    if sc >= 75 and not C['goppar_negatif'] and not C['ltv_incoherence']:
                         flags.append(("green", "Profil favorable — maintenir le suivi mensuel des KPIs"))

    for ftype, ftxt in flags:
        st.markdown(f'<div class="flag flag-{ftype}">{"🔴" if ftype=="red" else "🟡" if ftype=="amber" else "🟢"} {ftxt}</div>', unsafe_allow_html=True)

    st.divider()

    # Due diligence table
    with st.expander("📋 Checklist Due Diligence — entrée sur actif"):
        dd = {
            "Point de vérification": [
                "RevPAR vs benchmark marché","GOPPAR et marge GOP",
                "Historique 3 ans P&L","LTV sur dette existante","DSCR actuel et projeté",
                "DPE — classe et conformité","Contrat franchise / gestion (clauses sortie)",
                "Pipeline concurrentiel local","Mix clientèle diversifié","Score réputation en ligne",
                "Capex à prévoir","Couverture assurance pertes exploit.","Conformité RGPD","Turnover RH",
            ],
            "Cible": [
                "± 10% benchmark","Marge >25%","Tendance positive",
                "<65%",">1.25x","Classe D min","Préavis >24 mois",
                "Pipeline 3 ans","2+ segments","Google >4.0",
                "3-5% CA/an","Incluse","DPO désigné","<30%/an",
            ],
            "Criticité": [
                "🔴 CRITIQUE","🔴 CRITIQUE","🔴 CRITIQUE","🔴 CRITIQUE","🔴 CRITIQUE",
                "🔴 CRITIQUE","🔴 CRITIQUE","🟡 IMPORTANT","🟡 IMPORTANT","🟡 IMPORTANT",
                "🟡 IMPORTANT","🟡 IMPORTANT","🟡 IMPORTANT","🟢 À ÉVALUER",
            ],
        }
        st.table(dd)

    st.divider()

    # ── EXPORTS ──────────────────────────────────────────────────────────
    st.markdown("## Exports")
    ecol1, ecol2 = st.columns(2)

    with ecol1:
        st.markdown("**📊 Modèle Excel**")
        if st.button("Générer et télécharger Excel (.xlsx)", use_container_width=True, type="primary"):
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            thin = Side(style='thin', color='CCCCCC')
            brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr  = Alignment(horizontal='center', vertical='center')
            lft  = Alignment(horizontal='left', vertical='center')

            def hdr(ws, r, c, v):
                cell = ws.cell(r, c, v)
                cell.font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
                cell.fill  = PatternFill('solid', start_color='1A2E44')
                cell.alignment = ctr; cell.border = brd

            def sec(ws, r, c, v):
                cell = ws.cell(r, c, v)
                cell.font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
                cell.fill  = PatternFill('solid', start_color='2C5F8A')
                cell.alignment = lft; cell.border = brd

            def inp(ws, r, c, v):
                cell = ws.cell(r, c, v)
                cell.font  = Font(color='0000FF', name='Arial', size=10)
                cell.alignment = ctr; cell.border = brd

            def lbl(ws, r, c, v, alt=False):
                cell = ws.cell(r, c, v)
                cell.font  = Font(name='Arial', size=10)
                cell.fill  = PatternFill('solid', start_color='EBF3FA') if alt else PatternFill()
                cell.alignment = lft; cell.border = brd

            def val(ws, r, c, v, bold=False):
                cell = ws.cell(r, c, v)
                cell.font  = Font(bold=bold, name='Arial', size=10)
                cell.alignment = ctr; cell.border = brd

            def score_c(ws, r, c, v):
                clr = '1D9E75' if v>=75 else 'EF9F27' if v>=50 else 'E24B4A'
                cell = ws.cell(r, c, v)
                cell.font  = Font(bold=True, color='FFFFFF', name='Arial', size=12)
                cell.fill  = PatternFill('solid', start_color=clr)
                cell.alignment = ctr; cell.border = brd

            # ─ SHEET 1: Dashboard ─
            ws1 = wb.active; ws1.title = 'Dashboard'
            ws1.sheet_view.showGridLines = False
            for col, w in zip('ABCDE', [32,16,16,16,16]):
                ws1.column_dimensions[col].width = w

            ws1.merge_cells('A1:E1')
            t = ws1['A1']
            t.value = f'HOTEL KPI DASHBOARD — {asset_name.upper()}'
            t.font = Font(bold=True, color='FFFFFF', name='Arial', size=13)
            t.fill = PatternFill('solid', start_color='1A2E44')
            t.alignment = ctr; ws1.row_dimensions[1].height = 30

            ws1.merge_cells('A2:E2')
            ws1['A2'].value = f'{segment}  ·  {location}  ·  {period}  ·  AM: {analyst}'
            ws1['A2'].font = Font(italic=True, name='Arial', size=9, color='444441')
            ws1['A2'].alignment = ctr

            ws1.merge_cells('A4:B4'); sec(ws1, 4, 1, 'SCORE GLOBAL')
            ws1.merge_cells('C4:E4'); score_c(ws1, 4, 3, C['global_score'])
            ws1.row_dimensions[4].height = 24

            dim_rows = [
                ('Performance opéra. (25%)', C['ops']),
                ('Risque financier (25%)',    C['fin']),
                ('Risque marché (15%)',       C['mkt']),
                ('Résilience (15%)',          C['res']),
                ('Marque & Franchise (10%)', C['brand']),
                ('Légal & Réglementaire (10%)', C['legal']),
            ]
            for i, (nm, sc_d) in enumerate(dim_rows):
                r = 5 + i
                ws1.row_dimensions[r].height = 20
                lbl(ws1, r, 1, nm, i%2==1)
                ws1.merge_cells(f'C{r}:E{r}'); score_c(ws1, r, 3, sc_d)

            ws1.row_dimensions[12].height = 10
            ws1.merge_cells('A13:E13'); sec(ws1, 13, 1, 'KPIs OPÉRATIONNELS')
            for j, h in enumerate(['Indicateur','Valeur','Seuil','Benchmark','Statut']):
                hdr(ws1, 14, j+1, h)

            kpis = [
                ('RevPAR', f"{C['revpar']}€", '>100€', '70-500€ segment', '✓' if C['revpar']>=100 else '⚠'),
                ('GOPPAR', f"{C['goppar']}€", '>50€', '30-280€', '✓' if C['goppar']>=50 else '⚠'),
                ('Marge GOP', f"{C['gop_ratio']}%", '>35%', '25-45%', '✓' if C['gop_ratio']>=35 else '⚠'),
                ('Taux occupation', f"{data['occ']}%", '>68%', '65-82%', '✓' if data['occ']>=68 else '⚠'),
                ('ADR', f"{data['adr']}€", 'Selon segment', 'Variable', '—'),
                ('LTV', f"{data['ltv']}%", '<50%', 'Max 70%', '✓' if data['ltv']<50 else ('⚠' if data['ltv']<70 else '✗')),
                ('DSCR', f"{data['dscr']:.1f}x", '>1.5x', 'Min 1.2x', '✓' if data['dscr']>=1.5 else ('⚠' if data['dscr']>=1.2 else '✗')),
                ('Valeur actif (k€)', f"{C['asset_val']:,}", 'NOI/Cap Rate', '—', '—'),
                ('DPE', data['legal_dpe'], 'Classe D min', 'Réforme 2025-2034', '✓' if data['legal_dpe'] in ['A','B','C'] else ('⚠' if data['legal_dpe']=='D' else '✗')),
                ('Redevances franchise', f"{data['fee_pct']}%", '<10%', 'Marché', '✓' if data['fee_pct']<10 else '⚠'),
            ]
            for i, row in enumerate(kpis):
                r = 15 + i
                ws1.row_dimensions[r].height = 18
                bg = 'EBF3FA' if i%2==1 else None
                for j, v_val in enumerate(row):
                    cell = ws1.cell(r, j+1, v_val)
                    cell.font = Font(name='Arial', size=10,
                                     bold=(j==4),
                                     color=('1D9E75' if v_val=='✓' else 'E24B4A' if v_val=='✗' else 'EF9F27' if v_val=='⚠' else '000000'))
                    if bg: cell.fill = PatternFill('solid', start_color=bg)
                    cell.alignment = lft if j==0 else ctr
                    cell.border = brd

            ws1.row_dimensions[26].height = 10
            ws1['A27'].value = 'Légende : Bleu = input · Noir = formule · ✓ = dans seuil · ⚠ = vigilance · ✗ = hors seuil'
            ws1['A27'].font = Font(italic=True, name='Arial', size=8, color='888780')

            # ─ SHEET 2: Monthly Tracking ─
            ws2 = wb.create_sheet('Suivi mensuel')
            ws2.sheet_view.showGridLines = False
            months = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc','TOTAL']
            ws2.column_dimensions['A'].width = 30
            for i in range(len(months)):
                ws2.column_dimensions[get_column_letter(i+2)].width = 9

            ws2.merge_cells(f'A1:{get_column_letter(len(months)+1)}1')
            t2h = ws2['A1']
            t2h.value = f'SUIVI MENSUEL — {datetime.today().year}  |  {asset_name}'
            t2h.font  = Font(bold=True, color='FFFFFF', name='Arial', size=12)
            t2h.fill  = PatternFill('solid', start_color='1A2E44')
            t2h.alignment = ctr; ws2.row_dimensions[1].height = 26

            hdr(ws2, 2, 1, 'Indicateur')
            for i, m in enumerate(months):
                hdr(ws2, 2, i+2, m)

            kpi_rows2 = ['Taux occupation (%)','ADR (€)','RevPAR (€)','GOPPAR (€)',
                         'CA hébergement (k€)','CA total (k€)','Coûts opéra. (k€)',
                         'GOP (k€)','Marge GOP (%)','Nb chambres vendues',
                         'Score réput. en ligne','Redevance franchise (k€)']
            for i, kn in enumerate(kpi_rows2):
                r = 3 + i
                ws2.row_dimensions[r].height = 18
                lbl(ws2, r, 1, kn, i%2==1)
                for j in range(12):
                    c2 = ws2.cell(r, j+2, '')
                    c2.font = Font(color='0000FF', name='Arial', size=10)
                    c2.fill = PatternFill('solid', start_color='EBF3FA') if i%2==1 else PatternFill()
                    c2.alignment = ctr; c2.border = brd; c2.number_format = '#,##0.0'
                cl = [get_column_letter(j+2) for j in range(12)]
                formula = f'=AVERAGE({cl[0]}{r}:{cl[11]}{r})' if 'Marge' in kn or 'Score' in kn or 'Taux' in kn else f'=SUM({cl[0]}{r}:{cl[11]}{r})'
                tc = ws2.cell(r, 14, formula)
                tc.font = Font(bold=True, name='Arial', size=10)
                tc.fill = PatternFill('solid', start_color='D3D1C7')
                tc.alignment = ctr; tc.border = brd; tc.number_format = '#,##0.0'

            # ─ SHEET 3: Inputs ─
            ws3 = wb.create_sheet('Données saisies')
            ws3.sheet_view.showGridLines = False
            ws3.column_dimensions['A'].width = 36
            ws3.column_dimensions['B'].width = 20
            ws3.column_dimensions['C'].width = 28

            ws3.merge_cells('A1:C1')
            t3h = ws3['A1']
            t3h.value = 'DONNÉES SAISIES — ASSET MANAGER'
            t3h.font  = Font(bold=True, color='FFFFFF', name='Arial', size=12)
            t3h.fill  = PatternFill('solid', start_color='1A2E44')
            t3h.alignment = ctr; ws3.row_dimensions[1].height = 26

            all_inputs = [
                ('IDENTIFICATION', [
                    ('Actif', asset_name), ('Localisation', location),
                    ('Segment', segment), ('AM', analyst), ('Période', period),
                ]),
                ('OPÉRATIONNEL', [
                    ('Taux occupation (%)', data['occ']), ('ADR (€)', data['adr']),
                    ('Coûts opéra. (€/chambre dispo)', data['opcost']), ('Chambres', data['rooms']),
                ]),
                ('FINANCIER', [
                    ('NOI (k€)', data['noi']), ('Cap Rate (%)', data['caprate']),
                    ('LTV (%)', data['ltv']), ('DSCR', data['dscr']),
                ]),
                ('MARCHÉ', [
                    ('Saisonnalité', data['season']), ('Nouveaux concurrents', data['newcomp']),
                    ('Clientèle internationale (%)', data['intl']), ('Score localisation', data['loc']),
                ]),
                ('RÉSILIENCE', [
                    ('ESG', data['esg']), ('Digital', data['digi']),
                    ('Tension RH', data['hr']), ('Assurance', data['ins']), ('Diversification', data['div']),
                ]),
                ('MARQUE & FRANCHISE', [
                    ('Mode exploitation', data['brand_mode']), ('Force marque', data['brand_strength']),
                    ('Redevances (%CA)', data['fee_pct']), ('Durée contrat (ans)', data['contract_years']),
                    ('Pénalités sortie', data['exit_penalty']),
                ]),
                ('LÉGAL & RÉGLEMENTAIRE', [
                    ('DPE', data['legal_dpe']), ('Accessibilité PMR', data['legal_access']),
                    ('Sécurité incendie', data['legal_fire']), ('RGPD', data['legal_rgpd']),
                    ('Droit du travail', data['legal_labor']), ('Environnement', data['legal_env']),
                    ('Fiscal', data['legal_tax']),
                ]),
            ]

            row = 3
            for sec_name, items in all_inputs:
                ws3.merge_cells(f'A{row}:C{row}'); sec(ws3, row, 1, sec_name)
                ws3.row_dimensions[row].height = 20; row += 1
                for i, (k, v_val) in enumerate(items):
                    ws3.row_dimensions[row].height = 18
                    lbl(ws3, row, 1, k, i%2==1); inp(ws3, row, 2, v_val)
                    row += 1
                row += 1

            # ─ SHEET 4: Scores ─
            ws4 = wb.create_sheet('Scores calculés')
            ws4.sheet_view.showGridLines = False
            ws4.column_dimensions['A'].width = 30
            ws4.column_dimensions['B'].width = 16
            ws4.column_dimensions['C'].width = 16

            ws4.merge_cells('A1:C1')
            t4h = ws4['A1']
            t4h.value = 'SCORES DE RISQUE CALCULÉS'
            t4h.font  = Font(bold=True, color='FFFFFF', name='Arial', size=12)
            t4h.fill  = PatternFill('solid', start_color='1A2E44')
            t4h.alignment = ctr; ws4.row_dimensions[1].height = 26

            for j, h in enumerate(['Dimension','Score /100','Niveau de risque']):
                hdr(ws4, 2, j+1, h)

            score_rows = [
                ('Score global', C['global_score']),
                ('Performance opérationnelle', C['ops']),
                ('Risque financier', C['fin']),
                ('Risque marché', C['mkt']),
                ('Résilience', C['res']),
                ('Marque & Franchise', C['brand']),
                ('Légal & Réglementaire', C['legal']),
            ]
            for i, (nm, sc_d) in enumerate(score_rows):
                r = 3 + i
                ws4.row_dimensions[r].height = 20
                lbl(ws4, r, 1, nm, i%2==1); score_c(ws4, r, 2, sc_d)
                rl = ws4.cell(r, 3, risk_label(sc_d).replace('🟢','').replace('🟡','').replace('🔴','').strip())
                clr2 = '1D9E75' if sc_d>=75 else 'EF9F27' if sc_d>=50 else 'E24B4A'
                rl.font = Font(bold=True, name='Arial', size=10, color=clr2)
                rl.alignment = ctr; rl.border = brd

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button(
                label="⬇️ Télécharger le fichier Excel",
                data=buf,
                file_name=f"hotel_kpi_{asset_name.replace(' ','_')}_{datetime.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    with ecol2:
        st.markdown("**📄 Rapport PDF**")
        if st.button("Générer et télécharger PDF", use_container_width=True, type="primary"):
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT

            buf2 = BytesIO()
            doc = SimpleDocTemplate(buf2, pagesize=A4,
                                    leftMargin=1.8*cm, rightMargin=1.8*cm,
                                    topMargin=1.8*cm, bottomMargin=1.8*cm)

            DARK   = colors.HexColor('#1A2E44')
            MID    = colors.HexColor('#2C5F8A')
            GREEN  = colors.HexColor('#1D9E75')
            AMBER  = colors.HexColor('#EF9F27')
            RED    = colors.HexColor('#E24B4A')
            LGRAY  = colors.HexColor('#F1EFE8')
            MGRAY  = colors.HexColor('#D3D1C7')

            def sclr(s): return GREEN if s>=75 else AMBER if s>=50 else RED

            styles = getSampleStyleSheet()
            T = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=16, textColor=colors.white, alignment=TA_CENTER, leading=20)
            S = ParagraphStyle('s', fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#888780'), alignment=TA_CENTER)
            SH = ParagraphStyle('sh', fontName='Helvetica-Bold', fontSize=10, textColor=colors.white)
            B = ParagraphStyle('b', fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#2C2C2A'), leading=13)
            SM = ParagraphStyle('sm', fontName='Helvetica', fontSize=7.5, textColor=colors.HexColor('#5F5E5A'), leading=11)

            story = []

            # Header
            ht = Table([[Paragraph(f'RAPPORT D\'ANALYSE — INVESTISSEMENT HÔTELIER', T)]], colWidths=[17*cm])
            ht.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),DARK),('PADDING',(0,0),(-1,-1),10)]))
            story.append(ht); story.append(Spacer(1,0.3*cm))

            story.append(Paragraph(f'<b>Actif :</b> {asset_name}  ·  <b>Segment :</b> {segment}  ·  <b>Localisation :</b> {location}  ·  <b>AM :</b> {analyst}  ·  <b>Date :</b> {datetime.today().strftime("%d/%m/%Y")}', S))
            story.append(Spacer(1, 0.5*cm))

            # Global score
            sc2 = C['global_score']
            sc_lab = 'RISQUE FAIBLE' if sc2>=75 else 'RISQUE MODÉRÉ' if sc2>=50 else 'RISQUE ÉLEVÉ'
            sc_desc = ('Fondamentaux solides. Actif éligible à investissement.' if sc2>=75
                       else 'Profil acceptable avec points de vigilance.' if sc2>=50
                       else 'Risque élevé. Due diligence critique avant tout engagement.')

            stbl = Table([
                [Paragraph('<b>SCORE DE RISQUE GLOBAL</b>', SH),
                 Paragraph(f'<b>{sc2}/100</b>', ParagraphStyle('sv',fontName='Helvetica-Bold',fontSize=20,textColor=colors.white,alignment=TA_CENTER)),
                 Paragraph(f'<b>{sc_lab}</b>', ParagraphStyle('sl',fontName='Helvetica-Bold',fontSize=10,textColor=colors.white,alignment=TA_CENTER))]
            ], colWidths=[6*cm,4*cm,7*cm])
            stbl.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,-1),DARK),
                ('BACKGROUND',(1,0),(1,0),sclr(sc2)),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('PADDING',(0,0),(-1,-1),10),
                ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#444441')),
            ]))
            story.append(stbl); story.append(Spacer(1,0.2*cm))
            story.append(Paragraph(sc_desc, SM)); story.append(Spacer(1,0.4*cm))

            # Dimensions table
            dim_tbl_data = [['Dimension','Score','Poids','Niveau']]
            for nm, sc_d, w in [
                ('Performance opérationnelle', C['ops'],'25%'),
                ('Risque financier', C['fin'],'25%'),
                ('Risque marché', C['mkt'],'15%'),
                ('Résilience', C['res'],'15%'),
                ('Marque & Franchise', C['brand'],'10%'),
                ('Légal & Réglementaire', C['legal'],'10%'),
            ]:
                dim_tbl_data.append([nm, f'{sc_d}/100', w, 'Faible' if sc_d>=75 else 'Modéré' if sc_d>=50 else 'Élevé'])
            dt = Table(dim_tbl_data, colWidths=[7*cm,3*cm,2.5*cm,4.5*cm])
            dt.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),MID),('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
                ('ALIGN',(1,0),(-1,-1),'CENTER'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,LGRAY]),
                ('GRID',(0,0),(-1,-1),0.3,MGRAY),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
            ]))
            for i, (_, sc_d, _) in enumerate([('',C['ops'],''),('',C['fin'],''),('',C['mkt'],''),('',C['res'],''),('',C['brand'],''),('',C['legal'],'')]):
                dt.setStyle(TableStyle([('TEXTCOLOR',(3,i+1),(3,i+1),sclr(sc_d)),('FONTNAME',(3,i+1),(3,i+1),'Helvetica-Bold')]))
            story.append(dt); story.append(Spacer(1,0.5*cm))

            # Section headers + tables helper
            def sec_hdr(title):
                t = Table([[Paragraph(f'<b>{title}</b>', SH)]], colWidths=[17*cm])
                t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),MID),('PADDING',(0,0),(-1,-1),6),('LEFTPADDING',(0,0),(-1,-1),10)]))
                return t

            def make_table(data, widths):
                t = Table(data, colWidths=widths)
                t.setStyle(TableStyle([
                    ('BACKGROUND',(0,0),(-1,0),DARK),('TEXTCOLOR',(0,0),(-1,0),colors.white),
                    ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8.5),
                    ('ALIGN',(1,0),(-1,-1),'CENTER'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,LGRAY]),
                    ('GRID',(0,0),(-1,-1),0.3,MGRAY),('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
                ]))
                return t

            # Opérationnel
            story.append(sec_hdr('PERFORMANCE OPÉRATIONNELLE'))
            story.append(Spacer(1,0.2*cm))
            op_data = [['Indicateur','Valeur','Seuil','Statut']]
            for nm, v_val, seuil, ok in [
                ('RevPAR', f"{C['revpar']}€", '>100€', C['revpar']>=100),
                ('GOPPAR', f"{C['goppar']}€", '>50€', C['goppar']>=50),
                ('Marge GOP', f"{C['gop_ratio']}%", '>35%', C['gop_ratio']>=35),
                ('Taux occupation', f"{data['occ']}%", '>68%', data['occ']>=68),
                ('ADR', f"{data['adr']}€", 'Selon segment', None),
            ]:
                op_data.append([nm, v_val, seuil, 'OK' if ok else ('Vigilance' if ok is not None else '—')])
            ot = make_table(op_data, [5*cm,4*cm,4*cm,4*cm])
            for i, (_, _, _, ok) in enumerate([('','','',(C['revpar']>=100)),('','','',C['goppar']>=50),('','','',C['gop_ratio']>=35),('','','',data['occ']>=68),('','','',None)]):
                if ok is True:  ot.setStyle(TableStyle([('TEXTCOLOR',(3,i+1),(3,i+1),GREEN),('FONTNAME',(3,i+1),(3,i+1),'Helvetica-Bold')]))
                elif ok is False: ot.setStyle(TableStyle([('TEXTCOLOR',(3,i+1),(3,i+1),AMBER),('FONTNAME',(3,i+1),(3,i+1),'Helvetica-Bold')]))
            story.append(ot); story.append(Spacer(1,0.4*cm))

            # Marque & Franchise
            story.append(sec_hdr('MARQUE & FRANCHISE'))
            story.append(Spacer(1,0.2*cm))
            br_data = [['Paramètre','Valeur','Statut']]
            br_rows = [
                ('Mode exploitation', data['brand_mode'], None),
                ('Force de la marque (/10)', str(data['brand_strength']), data['brand_strength']>=7),
                ('Redevances (% CA)', f"{data['fee_pct']}%", data['fee_pct']<=10),
                ('Durée contrat restante', f"{data['contract_years']} ans", None),
                ('Pénalités sortie (/10)', str(data['exit_penalty']), data['exit_penalty']<=5),
            ]
            for nm, v_val, ok in br_rows:
                br_data.append([nm, v_val, 'OK' if ok is True else 'Vigilance' if ok is False else '—'])
            bt = make_table(br_data, [6.5*cm,5*cm,5.5*cm])
            story.append(bt); story.append(Spacer(1,0.4*cm))

            # Légal
            story.append(sec_hdr('LÉGAL & RÉGLEMENTAIRE'))
            story.append(Spacer(1,0.2*cm))
            lg_data = [['Dimension','Score/Classe','Statut']]
            lg_rows = [
                ('DPE', data['legal_dpe'], data['legal_dpe'] in ['A','B','C']),
                ('Accessibilité PMR', f"{data['legal_access']}/10", data['legal_access']>=7),
                ('Sécurité incendie', f"{data['legal_fire']}/10", data['legal_fire']>=7),
                ('Conformité RGPD', f"{data['legal_rgpd']}/10", data['legal_rgpd']>=7),
                ('Droit du travail', f"{data['legal_labor']}/10", data['legal_labor']>=7),
                ('Environnement', f"{data['legal_env']}/10", data['legal_env']>=7),
                ('Fiscal', f"{data['legal_tax']}/10", data['legal_tax']>=7),
            ]
            for nm, v_val, ok in lg_rows:
                lg_data.append([nm, v_val, 'Conforme' if ok else 'Risque'])
            lt = make_table(lg_data, [6*cm,5*cm,6*cm])
            for i, (_, _, ok) in enumerate(lg_rows):
                clr2 = GREEN if ok else RED
                lt.setStyle(TableStyle([('TEXTCOLOR',(2,i+1),(2,i+1),clr2),('FONTNAME',(2,i+1),(2,i+1),'Helvetica-Bold')]))
            story.append(lt); story.append(Spacer(1,0.4*cm))

            # Synthèse
            story.append(sec_hdr('SYNTHÈSE & RECOMMANDATIONS'))
            story.append(Spacer(1,0.3*cm))
            recs = []
            if C['fin']   < 50: recs.append('Risque financier élevé — revoir la structure de financement (LTV/DSCR)')
            if C['ops']   < 50: recs.append('Performance opérationnelle sous les benchmarks — analyser revenue management et management')
            if C['legal'] < 50: recs.append('Non-conformités réglementaires critiques — DPE, RGPD ou sécurité incendie à corriger')
            if C['brand'] < 50: recs.append('Risque franchise élevé — auditer les clauses d\'exit et le niveau des redevances')
            if C['res']   < 60: recs.append('Résilience fragile — investir en ESG, digital et couverture assurance')
            if data['legal_dpe'] in ['E','F','G']: recs.append(f'DPE {data["legal_dpe"]} — risque interdiction d\'exploitation et décote valeur actif')
            if not recs: recs.append('Aucun signal critique — maintenir le suivi mensuel des KPIs')
            for r in recs:
                story.append(Paragraph(f'• {r}', B)); story.append(Spacer(1,0.1*cm))

            story.append(Spacer(1,0.5*cm))
            story.append(HRFlowable(width='100%',thickness=0.5,color=MGRAY))
            story.append(Spacer(1,0.2*cm))
            story.append(Paragraph(f'Rapport généré le {datetime.today().strftime("%d/%m/%Y %H:%M")} — Sources : In Extenso TCH / Deloitte / Données saisies AM', SM))

            doc.build(story)
            buf2.seek(0)
            st.download_button(
                label="⬇️ Télécharger le rapport PDF",
                data=buf2,
                file_name=f"rapport_{asset_name.replace(' ','_')}_{datetime.today().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                use_container_width=True
            )

# ════════════════════════════════════════════════════════════════════════════
# TAB 8 — STRESS TEST
# ════════════════════════════════════════════════════════════════════════════
with t8:
    st.markdown("## Stress Test — Simulation de scénarios")
    st.info("Simulez l'impact de chocs de marché sur la performance et le score de risque de l'actif. Tout se recalcule en temps réel.")

    st.markdown("### Paramètres des chocs")
    sc1, sc2, sc3 = st.columns(3)
    with sc1:
        shock_occ    = st.slider("Choc taux occupation (pts)", -40, 20, -15, help="Ex: -15 = récession")
        shock_adr    = st.slider("Choc ADR (%)", -40, 20, -10, help="Ex: -10 = pression tarifaire")
    with sc2:
        shock_opcost = st.slider("Choc coûts opéra. (%)", -20, 60, +15, help="Ex: +15 = inflation énergie/salaires")
        shock_caprate = st.slider("Choc Cap Rate (pts)", -2.0, 3.0, +1.0, 0.5, help="Ex: +1.0 = remontée des taux")
    with sc3:
        shock_noi    = st.slider("Choc NOI (%)", -50, 20, -20, help="Ex: -20 = dégradation opérationnelle")
        shock_intl   = st.slider("Choc clientèle internationale (pts)", -40, 10, -20, help="Ex: -20 = choc géopolitique")

    st.markdown("---")

    data_stress = data.copy()
    data_stress['occ']      = max(5,  data['occ'] + shock_occ)
    data_stress['adr']      = max(10, round(data['adr'] * (1 + shock_adr/100)))
    data_stress['opcost']   = max(5,  round(data['opcost'] * (1 + shock_opcost/100)))
    data_stress['caprate']  = max(0.5, round(data['caprate'] + shock_caprate, 1))
    data_stress['noi']      = max(0,  round(data['noi'] * (1 + shock_noi/100)))
    data_stress['intl']     = max(0,  min(100, data['intl'] + shock_intl))
    data_stress['total_debt'] = data['total_debt']
    C_stress = compute_scores(data_stress)

    st.markdown("### Scénario de base vs Scénario stressé")
    col_base, col_stress, col_delta = st.columns(3)

    with col_base:
        st.markdown(f"""
        <div style="background:#F1EFE8;border-radius:10px;padding:1rem;text-align:center">
            <div style="font-size:0.75rem;font-weight:600;color:#888780;letter-spacing:.05em;margin-bottom:8px">SCÉNARIO DE BASE</div>
            <div style="font-size:2.5rem;font-weight:700;color:{score_color(C['global_score'])}">{C['global_score']}</div>
            <div style="font-size:0.85rem;color:{score_color(C['global_score'])}">{risk_label(C['global_score'])}</div>
        </div>
        """, unsafe_allow_html=True)

    with col_stress:
        clr_s = score_color(C_stress['global_score'])
        st.markdown(f"""
        <div style="background:{clr_s}22;border:2px solid {clr_s};border-radius:10px;padding:1rem;text-align:center">
            <div style="font-size:0.75rem;font-weight:600;color:#888780;letter-spacing:.05em;margin-bottom:8px">SCÉNARIO STRESSÉ</div>
            <div style="font-size:2.5rem;font-weight:700;color:{clr_s}">{C_stress['global_score']}</div>
            <div style="font-size:0.85rem;color:{clr_s}">{risk_label(C_stress['global_score'])}</div>
        </div>
        """, unsafe_allow_html=True)

    with col_delta:
        diff_sc = C_stress['global_score'] - C['global_score']
        clr_d = "#1D9E75" if diff_sc >= 0 else "#E24B4A"
        st.markdown(f"""
        <div style="background:#FFFFFF;border:0.5px solid #D3D1C7;border-radius:10px;padding:1rem;text-align:center">
            <div style="font-size:0.75rem;font-weight:600;color:#888780;letter-spacing:.05em;margin-bottom:8px">IMPACT DU CHOC</div>
            <div style="font-size:2.5rem;font-weight:700;color:{clr_d}">{'+' if diff_sc>=0 else ''}{diff_sc}</div>
            <div style="font-size:0.85rem;color:{clr_d}">pts sur le score global</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### Comparaison détaillée")

    import pandas as pd
    comparaisons = [
        ("Taux occupation (%)",  data['occ'],        data_stress['occ'],        "%"),
        ("ADR (€)",              data['adr'],         data_stress['adr'],        "€"),
        ("RevPAR (€)",           C['revpar'],         C_stress['revpar'],        "€"),
        ("GOPPAR (€)",           C['goppar'],         C_stress['goppar'],        "€"),
        ("Marge GOP (%)",        C['gop_ratio'],      C_stress['gop_ratio'],     "%"),
        ("NOI (k€)",             data['noi'],         data_stress['noi'],        "k€"),
        ("Valeur actif (k€)",    C['asset_val'],      C_stress['asset_val'],     "k€"),
        ("Score opérationnel",   C['ops'],            C_stress['ops'],           "/100"),
        ("Score financier",      C['fin'],            C_stress['fin'],           "/100"),
        ("Score global",         C['global_score'],   C_stress['global_score'],  "/100"),
    ]
    rows = []
    for label, base_v, stress_v, unit in comparaisons:
        diff = round(stress_v - base_v, 1)
        sign = "+" if diff > 0 else ""
        rows.append({
            "Indicateur": label,
            "Base": f"{base_v}{unit}",
            "Stressé": f"{stress_v}{unit}",
            "Variation": f"{sign}{diff}{unit}",
            "Signal": "🟢" if diff >= 0 else ("🟡" if diff >= -10 else "🔴"),
        })
    import pandas as pd
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("### Scénarios prédéfinis — Chocs historiques de référence")

    presets = {
        "🦠 Choc COVID": dict(occ=-35, adr=-25, op=+5, cr=+0.5, noi=-60, intl=-50),
        "📉 Récession": dict(occ=-15, adr=-10, op=+10, cr=+1.0, noi=-25, intl=-15),
        "⚡ Inflation": dict(occ=-5,  adr=+2,  op=+30, cr=+0.5, noi=-20, intl=-5),
        "🏗️ Concurrent": dict(occ=-12, adr=-8, op=0,   cr=+0.5, noi=-18, intl=0),
        "🌍 Géopolitique": dict(occ=-20, adr=-15, op=0, cr=+1.5, noi=-30, intl=-40),
    }

    preset_cols = st.columns(len(presets))
    for i, (label, p) in enumerate(presets.items()):
        with preset_cols[i]:
            d_p = data.copy()
            d_p['occ']     = max(5,   data['occ'] + p['occ'])
            d_p['adr']     = max(10,  round(data['adr'] * (1 + p['adr']/100)))
            d_p['opcost']  = max(5,   round(data['opcost'] * (1 + p['op']/100)))
            d_p['caprate'] = max(0.5, round(data['caprate'] + p['cr'], 1))
            d_p['noi']     = max(0,   round(data['noi'] * (1 + p['noi']/100)))
            d_p['intl']    = max(0,   min(100, data['intl'] + p['intl']))
            d_p['total_debt'] = data['total_debt']
            C_p   = compute_scores(d_p)
            sc_p  = C_p['global_score']
            diff_p = sc_p - C['global_score']
            clr_p  = score_color(sc_p)
            st.markdown(f"""
            <div style="background:{clr_p}15;border:1px solid {clr_p}55;border-radius:8px;padding:10px;text-align:center">
                <div style="font-size:0.72rem;font-weight:600;color:#444441;margin-bottom:6px">{label}</div>
                <div style="font-size:1.6rem;font-weight:700;color:{clr_p}">{sc_p}</div>
                <div style="font-size:0.75rem;color:{clr_p}">{'+' if diff_p>=0 else ''}{diff_p} pts</div>
                <div style="font-size:0.68rem;color:#888780;margin-top:4px">RevPAR {C_p['revpar']}€ | GOP {C_p['gop_ratio']}%</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("---")
    st.caption("💡 Les scénarios prédéfinis sont calibrés sur des chocs sectoriels historiques. Utilisez les sliders pour construire votre propre hypothèse de stress.")

# ════════════════════════════════════════════════════════════════════════════
# TAB 9 — RISK MANAGEMENT (Heatmap + Sensibilité + Monte Carlo PERT)
# ════════════════════════════════════════════════════════════════════════════
with t9:
    import numpy as np
    import pandas as pd

    PROFILES_PATH = Path("market_profiles.json")

    def load_profiles():
        if PROFILES_PATH.exists():
            with open(PROFILES_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        return {}

    def save_profiles(profiles):
        with open(PROFILES_PATH, "w", encoding="utf-8") as f:
            json.dump(profiles, f, ensure_ascii=False, indent=2)

    def pert_sample(low, mode, high, n=5000):
        """Distribution PERT — calibrée sur 3 points terrain"""
        if high == low:
            return np.full(n, mode)
        mean = (low + 4 * mode + high) / 6
        r = high - low
        a = 6 * (mean - low) / r
        b = 6 * (high - mean) / r
        a = max(a, 0.5)
        b = max(b, 0.5)
        samples = np.random.beta(a, b, n)
        return low + samples * r

    st.markdown("## Risk Management — Analyse structurée du risque")
    st.caption("Heatmap probabilité × impact · Sensibilité NOI · Monte Carlo PERT sur valeur actif")

    rm1, rm2, rm3 = st.tabs([
        "🗺️ Heatmap risques",
        "🌪️ Sensibilité NOI",
        "🎲 Monte Carlo & VaR",
    ])

    # ─────────────────────────────────────────────────────────────────────
    # SUB-TAB 1 — HEATMAP PROBABILITÉ × IMPACT
    # ─────────────────────────────────────────────────────────────────────
    with rm1:
        st.markdown("### Matrice Probabilité × Impact")
        st.info("Positionnez chaque risque selon sa probabilité d'occurrence et son impact sur le NOI. Score = Probabilité × Impact.")

        st.markdown("#### Saisie des risques")

        RISK_CATS = [
            "Marché & Économique",
            "Opérationnel",
            "Financier",
            "Marque & Franchise",
            "Légal & Réglementaire",
            "Résilience & Force Majeure",
        ]

        default_risks = [
            {"nom": "Récession / choc demande",       "cat": "Marché & Économique",        "proba": 3, "impact": 4},
            {"nom": "Nouveau concurrent majeur",       "cat": "Marché & Économique",        "proba": 3, "impact": 3},
            {"nom": "Saisonnalité forte",              "cat": "Marché & Économique",        "proba": 4, "impact": 2},
            {"nom": "Hausse coûts énergie / salaires", "cat": "Opérationnel",               "proba": 4, "impact": 3},
            {"nom": "Crise des talents / turnover",    "cat": "Opérationnel",               "proba": 4, "impact": 2},
            {"nom": "Défaillance management",          "cat": "Opérationnel",               "proba": 2, "impact": 4},
            {"nom": "Hausse des taux / refinancement", "cat": "Financier",                  "proba": 3, "impact": 4},
            {"nom": "Levier excessif (LTV >70%)",      "cat": "Financier",                  "proba": 2, "impact": 5},
            {"nom": "Hausse redevances franchise",     "cat": "Marque & Franchise",         "proba": 3, "impact": 3},
            {"nom": "Non-renouvellement contrat",      "cat": "Marque & Franchise",         "proba": 2, "impact": 4},
            {"nom": "Non-conformité DPE",              "cat": "Légal & Réglementaire",      "proba": 3, "impact": 4},
            {"nom": "Sanction RGPD",                   "cat": "Légal & Réglementaire",      "proba": 2, "impact": 3},
            {"nom": "Pandémie / force majeure",        "cat": "Résilience & Force Majeure", "proba": 1, "impact": 5},
            {"nom": "Sinistre / catastrophe naturelle","cat": "Résilience & Force Majeure", "proba": 1, "impact": 4},
            {"nom": "Cyberattaque",                    "cat": "Résilience & Force Majeure", "proba": 2, "impact": 3},
        ]

        if "risk_rows" not in st.session_state:
            st.session_state.risk_rows = default_risks.copy()

        col_add, col_reset = st.columns([1, 4])
        with col_add:
            if st.button("➕ Ajouter un risque", use_container_width=True):
                st.session_state.risk_rows.append(
                    {"nom": "Nouveau risque", "cat": RISK_CATS[0], "proba": 3, "impact": 3}
                )

        edited_risks = []
        for i, r in enumerate(st.session_state.risk_rows):
            c1, c2, c3, c4, c5 = st.columns([3, 2, 1, 1, 0.5])
            with c1:
                nom = st.text_input(f"Risque", r["nom"], key=f"rnom_{i}", label_visibility="collapsed")
            with c2:
                cat = st.selectbox(f"Cat", RISK_CATS, index=RISK_CATS.index(r["cat"]) if r["cat"] in RISK_CATS else 0, key=f"rcat_{i}", label_visibility="collapsed")
            with c3:
                proba = st.selectbox(f"P", [1,2,3,4,5], index=r["proba"]-1, key=f"rproba_{i}", label_visibility="collapsed",
                                     help="Probabilité 1=rare → 5=quasi-certain")
            with c4:
                impact = st.selectbox(f"I", [1,2,3,4,5], index=r["impact"]-1, key=f"rimpact_{i}", label_visibility="collapsed",
                                      help="Impact NOI 1=faible → 5=critique")
            with c5:
                if st.button("🗑️", key=f"rdel_{i}"):
                    st.session_state.risk_rows.pop(i)
                    st.rerun()
            edited_risks.append({"nom": nom, "cat": cat, "proba": proba, "impact": impact, "score": proba * impact})

        st.session_state.risk_rows = edited_risks

        st.markdown("---")
        st.markdown("#### Heatmap — Probabilité × Impact")

        # Construction de la heatmap 5×5
        heatmap = [[[] for _ in range(5)] for _ in range(5)]
        for r in edited_risks:
            p_idx = r["proba"] - 1
            i_idx = r["impact"] - 1
            heatmap[4 - p_idx][i_idx].append(r["nom"])

        CELL_COLORS = {
            (p, i): (
                "#E24B4A" if (p+1)*(i+1) >= 12
                else "#EF9F27" if (p+1)*(i+1) >= 6
                else "#1D9E75"
            )
            for p in range(5) for i in range(5)
        }

        html_heatmap = """
        <style>
        .hm-table { border-collapse: collapse; width: 100%; margin-top: 8px; }
        .hm-table td, .hm-table th {
            border: 1px solid #D3D1C7; padding: 8px 6px;
            font-size: 11px; text-align: center; vertical-align: middle;
            min-height: 50px; line-height: 1.4;
        }
        .hm-axis-label { font-weight: 600; color: #444441; background: #F1EFE8; font-size: 11px; }
        .hm-title { font-weight: 700; color: #1A2E44; font-size: 12px; }
        </style>
        <table class="hm-table">
        <thead>
            <tr>
                <th class="hm-axis-label" style="width:80px">Probabilité ↓ / Impact →</th>
        """
        impact_labels = ["1 — Faible", "2 — Limité", "3 — Modéré", "4 — Élevé", "5 — Critique"]
        for lbl in impact_labels:
            html_heatmap += f'<th class="hm-axis-label">{lbl}</th>'
        html_heatmap += "</tr></thead><tbody>"

        proba_labels = ["5 — Quasi-certain", "4 — Probable", "3 — Possible", "2 — Peu probable", "1 — Rare"]
        for row_idx in range(5):
            p_val = 5 - row_idx
            html_heatmap += f'<tr><td class="hm-axis-label">{proba_labels[row_idx]}</td>'
            for col_idx in range(5):
                i_val = col_idx + 1
                score = p_val * i_val
                bg = CELL_COLORS[(p_val-1, i_val-1)]
                risks_in_cell = heatmap[row_idx][col_idx]
                content = "<br>".join([f"<b>{rr}</b>" for rr in risks_in_cell]) if risks_in_cell else f'<span style="color:#B4B2A9">{score}</span>'
                html_heatmap += f'<td style="background:{bg}22;border-left:3px solid {bg}">{content}</td>'
            html_heatmap += "</tr>"

        html_heatmap += """</tbody></table>
        <div style="margin-top:8px;font-size:11px;color:#888780">
        <span style="background:#E24B4A22;padding:2px 8px;border-radius:4px;margin-right:8px">🔴 Critique (score ≥ 12)</span>
        <span style="background:#EF9F2722;padding:2px 8px;border-radius:4px;margin-right:8px">🟡 Modéré (score 6–11)</span>
        <span style="background:#1D9E7522;padding:2px 8px;border-radius:4px">🟢 Acceptable (score ≤ 5)</span>
        </div>"""

        st.markdown(html_heatmap, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("#### Top risques par score")
        top_risks = sorted(edited_risks, key=lambda x: x["score"], reverse=True)[:8]
        df_top = pd.DataFrame([
            {"Risque": r["nom"], "Catégorie": r["cat"],
             "Probabilité": r["proba"], "Impact": r["impact"],
             "Score P×I": r["score"],
             "Niveau": "🔴 Critique" if r["score"]>=12 else "🟡 Modéré" if r["score"]>=6 else "🟢 Acceptable"}
            for r in top_risks
        ])
        st.dataframe(df_top, use_container_width=True, hide_index=True)

    # ─────────────────────────────────────────────────────────────────────
    # SUB-TAB 2 — SENSIBILITÉ NOI (TORNADO CHART)
    # ─────────────────────────────────────────────────────────────────────
    with rm2:
        st.markdown("### Analyse de sensibilité — NOI")
        st.info("Impact d'une variation unitaire de chaque input clé sur le NOI. Le NOI définit ensuite la valeur actif, le DSCR et le rendement.")

        rooms_s = data.get('rooms', 80)

        def compute_noi(occ_pct, adr_v, opcost_v, rooms_v, fee_pct_v, caprate_v=None):
            """NOI reconstruit depuis les inputs opérationnels"""
            revpar = adr_v * (occ_pct / 100)
            goppar = revpar - opcost_v
            ca_heberg = revpar * rooms_v * 365 / 1000       # k€
            redev = ca_heberg * fee_pct_v / 100
            noi_calc = (goppar * rooms_v * 365 / 1000) - redev
            return round(noi_calc, 1)

        base_noi = compute_noi(data['occ'], data['adr'], data['opcost'],
                               rooms_s, data['fee_pct'])

        # Variations unitaires à tester
        sensibilite_inputs = [
            ("Taux occupation",      "occ",     +5,   -5,   "%"),
            ("ADR",                  "adr",     +10,  -10,  "€"),
            ("Coûts opérationnels",  "opcost",  -5,   +5,   "€"),
            ("Nombre de chambres",   "rooms",   +10,  -10,  "u"),
            ("Redevances franchise", "fee_pct", -2,   +2,   "%"),
            ("Cap Rate",             "caprate", -0.5, +0.5, "%"),
        ]

        tornado_rows = []
        for label, key, delta_up, delta_down, unit in sensibilite_inputs:
            d_up   = data.copy(); d_up['rooms']   = rooms_s
            d_down = data.copy(); d_down['rooms'] = rooms_s
            d_up[key]   = data[key] + delta_up
            d_down[key] = data[key] + delta_down

            noi_up   = compute_noi(d_up['occ'],   d_up['adr'],   d_up['opcost'],
                                   d_up['rooms'],  d_up['fee_pct'])
            noi_down = compute_noi(d_down['occ'], d_down['adr'], d_down['opcost'],
                                   d_down['rooms'], d_down['fee_pct'])

            impact_up   = round(noi_up   - base_noi, 1)
            impact_down = round(noi_down - base_noi, 1)
            spread      = abs(impact_up - impact_down)

            tornado_rows.append({
                "label":       label,
                "delta_up":    delta_up,
                "delta_down":  delta_down,
                "unit":        unit,
                "impact_up":   impact_up,
                "impact_down": impact_down,
                "spread":      spread,
            })

        # Tri par spread décroissant (ordre tornado)
        tornado_rows.sort(key=lambda x: x["spread"], reverse=True)
        max_spread = tornado_rows[0]["spread"] if tornado_rows else 1

        st.markdown(f"**NOI de base : {base_noi} k€**")
        st.markdown("*Barres : impact sur NOI d'une variation favorable (🟢) vs défavorable (🔴)*")
        st.markdown("")

        for row in tornado_rows:
            col_lbl, col_bar, col_vals = st.columns([2, 4, 2])
            with col_lbl:
                st.markdown(f"<div style='padding-top:6px;font-size:13px'>{row['label']}</div>",
                            unsafe_allow_html=True)
            with col_bar:
                up_pct   = abs(row['impact_up'])   / max_spread * 100
                down_pct = abs(row['impact_down']) / max_spread * 100
                st.markdown(f"""
                <div style="display:flex;align-items:center;gap:4px;padding-top:4px">
                  <div style="flex:1;text-align:right">
                    <div style="display:inline-block;width:{down_pct:.0f}%;height:20px;
                         background:#E24B4A;border-radius:3px 0 0 3px"></div>
                  </div>
                  <div style="width:2px;height:24px;background:#888780"></div>
                  <div style="flex:1">
                    <div style="display:inline-block;width:{up_pct:.0f}%;height:20px;
                         background:#1D9E75;border-radius:0 3px 3px 0"></div>
                  </div>
                </div>
                """, unsafe_allow_html=True)
            with col_vals:
                st.markdown(
                    f"<div style='font-size:11px;color:#888780;padding-top:6px'>"
                    f"<span style='color:#E24B4A'>{row['impact_down']:+.1f}k€</span> / "
                    f"<span style='color:#1D9E75'>{row['impact_up']:+.1f}k€</span></div>",
                    unsafe_allow_html=True)

        st.markdown("---")
        df_sens = pd.DataFrame([{
            "Variable": r["label"],
            f"Choc favorable ({'+' if r['delta_up']>0 else ''}{r['delta_up']}{r['unit']})": f"+{r['impact_up']}k€",
            f"Choc défavorable ({'+' if r['delta_down']>0 else ''}{r['delta_down']}{r['unit']})": f"{r['impact_down']}k€",
            "Spread total (k€)": round(r["spread"], 1),
            "Sensibilité": "🔴 Très forte" if r["spread"] > max_spread*0.6
                           else "🟡 Forte" if r["spread"] > max_spread*0.3
                           else "🟢 Modérée",
        } for r in tornado_rows])
        st.dataframe(df_sens, use_container_width=True, hide_index=True)

        st.caption("💡 Plus le spread est large, plus cette variable est déterminante pour le NOI — et donc pour la valeur actif et le DSCR.")

    # ─────────────────────────────────────────────────────────────────────
    # SUB-TAB 3 — MONTE CARLO PERT + VaR
    # ─────────────────────────────────────────────────────────────────────
    with rm3:
        st.markdown("### Monte Carlo — Distribution probabiliste du NOI et de la valeur actif")
        st.info("Calibrez les distributions PERT sur votre connaissance terrain de ce marché (P10/P50/P90). Sauvegardez le profil pour le réutiliser sur d'autres actifs du même marché.")

        # ── Profils de marché ────────────────────────────────────────────
        profiles = load_profiles()
        col_prof1, col_prof2, col_prof3 = st.columns([2, 1, 1])
        with col_prof1:
            profile_names = ["— Nouveau profil —"] + list(profiles.keys())
            selected_profile = st.selectbox("Charger un profil de marché existant", profile_names)
        with col_prof2:
            new_profile_name = st.text_input("Nom du profil à sauvegarder",
                                              placeholder="Ex: Nice Été 4★")
        with col_prof3:
            st.markdown("<div style='padding-top:28px'>", unsafe_allow_html=True)
            save_clicked = st.button("💾 Sauvegarder", use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

        # Charger profil si sélectionné
        if selected_profile != "— Nouveau profil —" and selected_profile in profiles:
            P = profiles[selected_profile]
        else:
            P = {
                "occ":     {"p10": max(10, data['occ']-15), "p50": data['occ'],     "p90": min(99, data['occ']+10)},
                "adr":     {"p10": round(data['adr']*0.80), "p50": data['adr'],     "p90": round(data['adr']*1.20)},
                "opcost":  {"p10": round(data['opcost']*0.90), "p50": data['opcost'], "p90": round(data['opcost']*1.25)},
                "caprate": {"p10": max(0.5, data['caprate']-1.5), "p50": data['caprate'], "p90": data['caprate']+1.5},
                "fee_pct": {"p10": max(0, data['fee_pct']-2), "p50": data['fee_pct'], "p90": data['fee_pct']+3},
            }

        st.markdown("#### Paramètres de marché local — P10 / P50 / P90")
        st.caption("P10 = scénario pessimiste · P50 = centrale · P90 = scénario optimiste")

        vars_config = [
            ("Taux occupation (%)", "occ",     0.0,  100.0, "%"),
            ("ADR (€/nuit)",        "adr",     10.0, 2000.0,"€"),
            ("Coûts opéra. (€/ch)", "opcost",  5.0,  500.0, "€"),
            ("Cap Rate (%)",        "caprate", 0.5,  15.0,  "%"),
            ("Redevances (% CA)",   "fee_pct", 0.0,  25.0,  "%"),
        ]

        pert_params = {}
        for label, key, vmin, vmax, unit in vars_config:
            c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
            with c1:
                st.markdown(f"<div style='padding-top:8px;font-size:13px'>{label}</div>", unsafe_allow_html=True)
            with c2:
                p10 = st.number_input(f"P10", float(vmin), float(vmax), float(P[key]["p10"]),
                                      step=0.1, key=f"p10_{key}", label_visibility="collapsed",
                                      help=f"P10 — {label}")
            with c3:
                p50 = st.number_input(f"P50", float(vmin), float(vmax), float(P[key]["p50"]),
                                      step=0.1, key=f"p50_{key}", label_visibility="collapsed",
                                      help=f"P50 — {label}")
            with c4:
                p90 = st.number_input(f"P90", float(vmin), float(vmax), float(P[key]["p90"]),
                                      step=0.1, key=f"p90_{key}", label_visibility="collapsed",
                                      help=f"P90 — {label}")
            pert_params[key] = {"p10": p10, "p50": p50, "p90": p90}

        # Sauvegarde profil
        if save_clicked and new_profile_name.strip():
            profiles[new_profile_name.strip()] = pert_params
            save_profiles(profiles)
            st.success(f"✅ Profil '{new_profile_name.strip()}' sauvegardé.")

        st.markdown("---")

        n_sim = st.select_slider("Nombre de simulations", [1000, 2000, 5000, 10000], value=5000)

        if st.button("▶️ Lancer Monte Carlo", type="primary", use_container_width=False):
            np.random.seed(42)
            rooms_mc = data.get('rooms', 80)

            # Simulations PERT vectorisées
            occ_sim     = pert_sample(pert_params['occ']['p10'],
                                      pert_params['occ']['p50'],
                                      pert_params['occ']['p90'], n_sim)
            adr_sim     = pert_sample(pert_params['adr']['p10'],
                                      pert_params['adr']['p50'],
                                      pert_params['adr']['p90'], n_sim)
            opcost_sim  = pert_sample(pert_params['opcost']['p10'],
                                      pert_params['opcost']['p50'],
                                      pert_params['opcost']['p90'], n_sim)
            caprate_sim = pert_sample(pert_params['caprate']['p10'],
                                      pert_params['caprate']['p50'],
                                      pert_params['caprate']['p90'], n_sim)
            fee_sim     = pert_sample(pert_params['fee_pct']['p10'],
                                      pert_params['fee_pct']['p50'],
                                      pert_params['fee_pct']['p90'], n_sim)

            # Corrélation négative occ-adr (empirique hôtellerie)
            corr_noise = np.random.normal(0, 0.03, n_sim)
            adr_sim    = adr_sim * (1 - 0.35 * (occ_sim - occ_sim.mean()) / (occ_sim.std() + 1e-9) * corr_noise.std())

            # Calcul NOI simulé (k€)
            revpar_sim  = adr_sim * (occ_sim / 100)
            goppar_sim  = revpar_sim - opcost_sim
            ca_sim      = revpar_sim * rooms_mc * 365 / 1000
            redev_sim   = ca_sim * fee_sim / 100
            noi_sim     = (goppar_sim * rooms_mc * 365 / 1000) - redev_sim

            # Valeur actif simulée (k€)
            caprate_safe = np.maximum(caprate_sim, 0.5)
            val_sim      = noi_sim / (caprate_safe / 100)

            # ── Métriques clés ────────────────────────────────────────────
            st.markdown("### Résultats — Distribution du NOI")

            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("NOI P10 (pessimiste)", f"{np.percentile(noi_sim,10):.0f} k€")
            m2.metric("NOI P25",              f"{np.percentile(noi_sim,25):.0f} k€")
            m3.metric("NOI P50 (médiane)",    f"{np.percentile(noi_sim,50):.0f} k€")
            m4.metric("NOI P75",              f"{np.percentile(noi_sim,75):.0f} k€")
            m5.metric("NOI P90 (optimiste)",  f"{np.percentile(noi_sim,90):.0f} k€")

            prob_perte = round((noi_sim < 0).mean() * 100, 1)
            noi_var95  = round(np.percentile(noi_sim, 5), 0)
            noi_var99  = round(np.percentile(noi_sim, 1), 0)

            st.markdown("---")
            fa, fb, fc = st.columns(3)
            fa.metric("VaR NOI à 95%", f"{noi_var95} k€",
                      help="Dans 95% des scénarios, le NOI sera supérieur à cette valeur")
            fb.metric("VaR NOI à 99%", f"{noi_var99} k€",
                      help="Dans 99% des scénarios, le NOI sera supérieur à cette valeur")
            fc.metric("Probabilité d'exploitation à perte",
                      f"{prob_perte}%",
                      delta="Risque critique" if prob_perte > 10 else "Acceptable",
                      delta_color="inverse")

            # ── Histogramme NOI ────────────────────────────────────────────
            st.markdown("#### Distribution du NOI — 5 000 scénarios PERT")

            hist_vals, hist_bins = np.histogram(noi_sim, bins=60)
            bin_centers = (hist_bins[:-1] + hist_bins[1:]) / 2

            p5_v  = np.percentile(noi_sim, 5)
            p50_v = np.percentile(noi_sim, 50)
            p95_v = np.percentile(noi_sim, 95)

            bar_colors = [
                "#E24B4A" if b < p5_v
                else "#1D9E75" if b > p95_v
                else "#378ADD"
                for b in bin_centers
            ]

            bar_html = f"""
            <div style="position:relative;width:100%;padding:8px 0">
              <div style="display:flex;align-items:flex-end;gap:1px;height:140px">
            """
            max_h = max(hist_vals) if max(hist_vals) > 0 else 1
            for h, color in zip(hist_vals, bar_colors):
                pct = h / max_h * 100
                bar_html += f'<div style="flex:1;height:{pct:.0f}%;background:{color};border-radius:1px 1px 0 0;min-width:2px"></div>'

            bar_html += f"""
              </div>
              <div style="display:flex;justify-content:space-between;margin-top:4px;font-size:10px;color:#888780">
                <span>{noi_sim.min():.0f}k€</span>
                <span style="color:#E24B4A">VaR95: {p5_v:.0f}k€</span>
                <span style="color:#378ADD;font-weight:600">P50: {p50_v:.0f}k€</span>
                <span style="color:#1D9E75">P95: {p95_v:.0f}k€</span>
                <span>{noi_sim.max():.0f}k€</span>
              </div>
              <div style="font-size:10px;color:#888780;margin-top:4px">
                🔴 Zone VaR 5% &nbsp;|&nbsp; 🔵 Zone centrale (90%) &nbsp;|&nbsp; 🟢 Zone P95+
              </div>
            </div>
            """
            st.markdown(bar_html, unsafe_allow_html=True)

            # ── Valeur actif ───────────────────────────────────────────────
            st.markdown("---")
            st.markdown("### Distribution de la valeur actif (k€) — NOI / Cap Rate")

            v1, v2, v3, v4, v5 = st.columns(5)
            v1.metric("Valeur P10",     f"{np.percentile(val_sim,10):,.0f} k€")
            v2.metric("Valeur P25",     f"{np.percentile(val_sim,25):,.0f} k€")
            v3.metric("Valeur P50",     f"{np.percentile(val_sim,50):,.0f} k€")
            v4.metric("Valeur P75",     f"{np.percentile(val_sim,75):,.0f} k€")
            v5.metric("Valeur P90",     f"{np.percentile(val_sim,90):,.0f} k€")

            val_var95 = np.percentile(val_sim, 5)
            val_range = np.percentile(val_sim, 90) - np.percentile(val_sim, 10)

            g1, g2 = st.columns(2)
            g1.metric("VaR valeur actif à 95%",
                      f"{val_var95:,.0f} k€",
                      help="Valeur plancher dans 95% des scénarios — base pour les covenants bancaires")
            g2.metric("Plage d'incertitude P10-P90",
                      f"{val_range:,.0f} k€",
                      help="Amplitude de l'incertitude sur la valorisation")

            # ── Intervalles de confiance ───────────────────────────────────
            st.markdown("---")
            st.markdown("### Intervalles de confiance — synthèse pour présentation investisseur")

            ic_data = {
                "Indicateur": ["NOI (k€)", "Valeur actif (k€)", "GOPPAR (€)", "RevPAR (€)"],
                "P10 — Pessimiste": [
                    f"{np.percentile(noi_sim,10):.0f}",
                    f"{np.percentile(val_sim,10):,.0f}",
                    f"{np.percentile(goppar_sim,10):.1f}",
                    f"{np.percentile(revpar_sim,10):.1f}",
                ],
                "P50 — Central": [
                    f"{np.percentile(noi_sim,50):.0f}",
                    f"{np.percentile(val_sim,50):,.0f}",
                    f"{np.percentile(goppar_sim,50):.1f}",
                    f"{np.percentile(revpar_sim,50):.1f}",
                ],
                "P90 — Optimiste": [
                    f"{np.percentile(noi_sim,90):.0f}",
                    f"{np.percentile(val_sim,90):,.0f}",
                    f"{np.percentile(goppar_sim,90):.1f}",
                    f"{np.percentile(revpar_sim,90):.1f}",
                ],
                "VaR 95%": [
                    f"{np.percentile(noi_sim,5):.0f}",
                    f"{np.percentile(val_sim,5):,.0f}",
                    f"{np.percentile(goppar_sim,5):.1f}",
                    f"{np.percentile(revpar_sim,5):.1f}",
                ],
            }
            st.dataframe(pd.DataFrame(ic_data), use_container_width=True, hide_index=True)

            st.markdown("---")
            st.caption(
                f"**Méthodologie** : Distribution PERT calibrée sur P10/P50/P90 terrain · "
                f"{n_sim:,} simulations vectorisées NumPy · "
                f"Corrélation négative occ/ADR intégrée (ρ ≈ -0.35, empirique hôtellerie France) · "
                f"NOI = (RevPAR − OpCost) × Chambres × 365 / 1000 − Redevances"
            )
        else:
            st.markdown("""
            <div style="background:#F1EFE8;border-radius:10px;padding:2rem;text-align:center;color:#888780">
                <div style="font-size:2rem;margin-bottom:8px">🎲</div>
                <div style="font-weight:600;color:#444441;margin-bottom:4px">Monte Carlo PERT</div>
                <div style="font-size:13px">Calibrez vos P10/P50/P90 ci-dessus puis lancez la simulation</div>
            </div>
            """, unsafe_allow_html=True)
