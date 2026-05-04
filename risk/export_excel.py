from io import BytesIO
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel_bytes(asset_name: str, inputs: Dict[str, Any], C: Dict[str, Any]) -> BytesIO:
    """Generate a compact Excel report and return it as BytesIO.

    This function is intentionally focused and testable. It mirrors the important
    pieces of the original app but keeps formatting minimal to remain robust.
    """
    wb = Workbook()
    thin = Side(style='thin', color='CCCCCC')
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal='center', vertical='center')
    lft = Alignment(horizontal='left', vertical='center')

    def hdr(ws, r, c, v):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = PatternFill('solid', start_color='1A2E44')
        cell.alignment = ctr
        cell.border = brd

    def lbl(ws, r, c, v, alt=False):
        cell = ws.cell(r, c, v)
        cell.font = Font(name='Arial', size=10)
        if alt:
            cell.fill = PatternFill('solid', start_color='EBF3FA')
        cell.alignment = lft
        cell.border = brd

    def val(ws, r, c, v, bold=False):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=bold, name='Arial', size=10)
        cell.alignment = ctr
        cell.border = brd

    # Sheet 1: Dashboard
    ws = wb.active
    ws.title = 'Dashboard'
    ws.sheet_view.showGridLines = False
    for col, w in zip(range(1, 7), [40, 16, 16, 16, 16, 16]):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells('A1:F1')
    t = ws['A1']
    t.value = f'HOTEL KPI DASHBOARD — {asset_name}'
    t.font = Font(bold=True, color='FFFFFF', name='Arial', size=13)
    t.fill = PatternFill('solid', start_color='1A2E44')
    t.alignment = ctr

    ws['A3'] = 'Score global'
    val(ws, 3, 2, C.get('global_score', '—'), bold=True)

    # KPIs
    ws['A5'] = 'KPIs calculés'
    hdr(ws, 6, 1, 'Indicateur')
    hdr(ws, 6, 2, 'Valeur')

    kpis = [
        ('RevPAR', f"{C.get('revpar','—')}€"),
        ('GOPPAR', f"{C.get('goppar','—')}€"),
        ('Marge GOP', f"{C.get('gop_ratio','—')}%"),
        ('LTV', f"{inputs.get('ltv','—')}%"),
        ('DSCR', f"{inputs.get('dscr','—')}x"),
        ('Valeur actif', f"{C.get('asset_val','—')}")
    ]

    for i, (k, v) in enumerate(kpis):
        r = 7 + i
        lbl(ws, r, 1, k, i % 2 == 1)
        val(ws, r, 2, v)

    # Sheet 2: Inputs
    ws2 = wb.create_sheet('Données saisies')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 36
    ws2.column_dimensions['B'].width = 20

    ws2.merge_cells('A1:B1')
    ws2['A1'] = 'DONNÉES SAISIES'
    ws2['A1'].font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
    ws2['A1'].fill = PatternFill('solid', start_color='1A2E44')
    ws2['A1'].alignment = ctr

    row = 3
    for k, v in sorted(inputs.items()):
        lbl(ws2, row, 1, str(k))
        val(ws2, row, 2, str(v))
        row += 1

    # Sheet 3: Scores
    ws3 = wb.create_sheet('Scores')
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 16

    hdr(ws3, 1, 1, 'Dimension')
    hdr(ws3, 1, 2, 'Score')

    score_rows = [
        ('Score global', C.get('global_score')),
        ('Opérationnel', C.get('ops')),
        ('Financier', C.get('fin')),
        ('Marché', C.get('mkt')),
        ('Résilience', C.get('res')),
        ('Marque', C.get('brand')),
        ('Légal', C.get('legal')),
    ]
    for i, (n, s) in enumerate(score_rows):
        r = 2 + i
        lbl(ws3, r, 1, n, i % 2 == 1)
        val(ws3, r, 2, s)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
